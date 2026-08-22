"""
=============================================================================
EXPORT 011 TABLES TO EXCEL — SISTEKIN UWG 2026
=============================================================================
Script ini membaca file Markdown:
  011_IMPLEMENTASI_OBE_SISTEKIN2026_TABLES.md
dan mengonversinya menjadi Workbook Excel multi-tab dengan formatting profesional:
  011_IMPLEMENTASI_OBE_SISTEKIN2026_TABLES.xlsx

Setiap section '## Sheet X: ...' akan menjadi 1 tab terpisah.
=============================================================================
"""

import sys
import io
import os
import re

if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] Library 'openpyxl' belum terinstall.")
    print("Silakan jalankan: pip install openpyxl")
    sys.exit(1)

# Direktori kerja berbasis lokasi script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

SRC_MD = os.path.join(PARENT_DIR, "011_IMPLEMENTASI_OBE_SISTEKIN2026_TABLES.md")
DEST_XLSX = os.path.join(PARENT_DIR, "011_IMPLEMENTASI_OBE_SISTEKIN2026_TABLES.xlsx")

# Palet Warna Profesional
COLOR = {
    'sheet_bg': '1F3864',    # Navy Biru Gelap
    'sheet_fg': 'FFFFFF',
    'col_header': '2E75B6',  # Biru Header Tabel
    'col_fg': 'FFFFFF',
    'section': 'BDD7EE',     # Biru Muda Section Header (A., B., C.)
    'section_fg': '1F3864',
    'alt_row': 'EEF4FB',     # Zebra Stripe
    'border': 'B8CCE4',
    'note': 'FFF2CC',        # Kuning Lembut (Summary/Catatan)
}

thin = Side(style='thin', color=COLOR['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style_header(cell, fg='FFFFFF', bg='2E75B6', bold=True, sz=10, wrap=True):
    cell.font = Font(bold=bold, color=fg, size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='center', horizontal='center')
    cell.border = BORDER

def style_data(cell, bg='FFFFFF', bold=False, sz=10, halign='left', wrap=True):
    cell.font = Font(bold=bold, color='000000', size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='top', horizontal=halign)
    cell.border = BORDER

def style_section(cell):
    cell.font = Font(bold=True, color=COLOR['section_fg'], size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['section'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_note(cell):
    cell.font = Font(bold=True, color='7F6000', size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['note'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_title(ws, title, subtitle, row_offset=1):
    ws.row_dimensions[row_offset].height = 24
    ws.row_dimensions[row_offset + 1].height = 16
    c1 = ws.cell(row=row_offset, column=1, value=title)
    c1.font = Font(bold=True, color=COLOR['sheet_fg'], size=13, name='Calibri')
    c1.fill = cell_fill(COLOR['sheet_bg'])
    c1.alignment = Alignment(vertical='center', horizontal='left')

    c2 = ws.cell(row=row_offset + 1, column=1, value=subtitle)
    c2.font = Font(bold=False, color=COLOR['sheet_fg'], size=10, name='Calibri')
    c2.fill = cell_fill(COLOR['sheet_bg'])
    c2.alignment = Alignment(vertical='center', horizontal='left')

def parse_md(path):
    if not os.path.exists(path):
        raise FileNotFoundError(f"File markdown tidak ditemukan: {path}")

    with open(path, encoding='utf-8') as f:
        raw = f.read()

    sheets = []
    parts = re.split(r'\n## (Sheet \d+:[^\n]+)', raw)
    for i in range(1, len(parts), 2):
        title_raw = parts[i].strip()
        body = parts[i + 1] if i + 1 < len(parts) else ''

        sub_match = re.search(r'^>\s*(.+)', body, re.MULTILINE)
        subtitle = sub_match.group(1).strip() if sub_match else ''

        tables = []
        table_blocks = re.findall(
            r'(\|.+\|\n\|[-| :]+\|\n(?:\|.+\|\n?)*)',
            body
        )
        for block in table_blocks:
            rows = []
            for line in block.strip().split('\n'):
                if re.match(r'\|[-| :]+\|', line):
                    continue
                cells = [c.strip() for c in line.strip('|').split('|')]
                rows.append(cells)
            if rows:
                tables.append({'rows': rows})

        sheets.append({
            'sheet_num': len(sheets) + 1,
            'title': title_raw,
            'subtitle': subtitle,
            'tables': tables,
        })

    return sheets

SECTION_RE = re.compile(r'^[A-F]\.\s')
NOTE_RE = re.compile(
    r'^(Jumlah|Validasi|Ringkasan|Target|%\s*Mahasiswa|TOTAL|rata|total)',
    re.IGNORECASE
)

def is_section_row(row):
    return any(SECTION_RE.match(c) for c in row)

def is_note_row(row):
    first = row[0] if row else ''
    return bool(NOTE_RE.match(first))

def write_sheet(ws, sheet_info):
    title = f"SISTEKIN 2026 — {sheet_info['title']}"
    subtitle = sheet_info['subtitle']

    style_title(ws, title, subtitle, row_offset=1)
    ws.freeze_panes = 'A4'

    current_row = 3

    for tbl_idx, tbl in enumerate(sheet_info['tables']):
        rows = tbl['rows']
        if not rows:
            continue

        header = rows[0]
        n_cols = len(header)

        if tbl_idx == 0:
            try:
                ws.merge_cells(
                    start_row=1, start_column=1,
                    end_row=1, end_column=max(n_cols, 6)
                )
                ws.merge_cells(
                    start_row=2, start_column=1,
                    end_row=2, end_column=max(n_cols, 6)
                )
            except Exception:
                pass

        ws.row_dimensions[current_row].height = 28
        for col_i, val in enumerate(header, start=1):
            c = ws.cell(row=current_row, column=col_i, value=val)
            style_header(c, sz=10)
        current_row += 1

        for row_i, row in enumerate(rows[1:], start=1):
            row = (row + [''] * n_cols)[:n_cols]

            if is_section_row(row):
                bg = COLOR['section']
                styler = style_section
                bold = True
            elif is_note_row(row):
                bg = COLOR['note']
                styler = style_note
                bold = True
            else:
                bg = COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF'
                styler = None
                bold = False

            ws.row_dimensions[current_row].height = 16

            for col_i, val in enumerate(row, start=1):
                c = ws.cell(row=current_row, column=col_i, value=val)
                if styler:
                    styler(c)
                else:
                    halign = 'center' if col_i <= 5 else 'left'
                    style_data(c, bg=bg, bold=bold, halign=halign)

            current_row += 1

        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            except Exception:
                pass
        adjusted = max(10, max_len * 1.15)
        ws.column_dimensions[col_letter].width = adjusted

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14

def convert():
    print(f"[INFO] Membaca sumber Markdown:\n  -> {SRC_MD}")
    sheets_data = parse_md(SRC_MD)
    print(f"[INFO] Ditemukan {len(sheets_data)} sheet.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cover = wb.create_sheet("Cover")
    cover.sheet_properties.tabColor = "1F3864"
    cover.column_dimensions['A'].width = 5
    cover.column_dimensions['B'].width = 85

    cover_rows = [
        (1, "IMPLEMENTASI MODUL OBE", True, 20, '1F3864', 'FFFFFF'),
        (2, "S1 SISTEM DAN TEKNOLOGI INFORMASI 2026", True, 14, '1F3864', 'FFFFFF'),
        (3, "FSTI — Universitas Widyagama Malang", False, 12, '2E75B6', 'FFFFFF'),
        (4, "", False, 10, 'FFFFFF', '000000'),
        (5, "Versi: Revisi Final 2026 (Kurikulum KPT-OBE SISTEKIN)", False, 11, 'EEF4FB', '1F3864'),
        (6, "Standar: Permendikbudristek No. 53/2023 | APTIKOM IS2020/IT2017 | LAM INFOKOM", False, 11, 'EEF4FB', '1F3864'),
        (7, "Total Sheet: 14 (Cover + Sheet 1–14)", False, 11, 'EEF4FB', '1F3864'),
        (8, "", False, 10, 'FFFFFF', '000000'),
        (9, "Parameter Kurikulum Final Terverifikasi:", True, 11, 'FFF2CC', '7F6000'),
        (10, "  ✓  Profil Lulusan: 4 PL (PL-1 s.d. PL-4)", False, 11, 'FFF2CC', '7F6000'),
        (11, "  ✓  CPL: 14 (S1, KU1-3, P1-4, KK1-6)", False, 11, 'FFF2CC', '7F6000'),
        (12, "  ✓  Paket Ditempuh: 55 MK / 146 SKS", False, 11, 'FFF2CC', '7F6000'),
        (13, "  ✓  Portofolio Ditawarkan: 67 MK / 182 SKS", False, 11, 'FFF2CC', '7F6000'),
        (15, "  ✓  STI-405 Dasar Keamanan Informasi = 2 SKS & STI-602 Smart City = 2 SKS", False, 11, 'FFF2CC', '7F6000'),
        (16, "  ✓  MKU-406 Agama II & MKU-508 Kewirausahaan II = 0 SKS (Kebijakan UWG)", False, 11, 'FFF2CC', '7F6000'),
        (17, "  ✓  Skema 4x Asesmen Baku per MK (Tugas 1 [20%], UTS [25-30%], Tugas 2 [20-25%], UAS [30%])", False, 11, 'FFF2CC', '7F6000'),
    ]

    for r, val, bold, sz, bg, fg in cover_rows:
        cover.row_dimensions[r].height = sz * 1.8
        c = cover.cell(row=r, column=2, value=val)
        c.font = Font(bold=bold, size=sz, color=fg, name='Calibri')
        c.fill = cell_fill(bg)
        c.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        ca = cover.cell(row=r, column=1, value='')
        ca.fill = cell_fill(bg)

    tab_colors = {
        1: '70AD47', 2: 'ED7D31', 3: '4472C4', 4: '4472C4',
        5: '4472C4', 6: '5B9BD5', 7: '5B9BD5', 8: '9DC3E6',
        9: 'FF0000', 10: 'FF0000', 11: 'FF0000', 12: 'FF0000',
        13: 'FFC000', 14: '7030A0',
    }

    for sd in sheets_data:
        snum = sd['sheet_num']
        raw_title = sd['title']
        tab_name = re.sub(r'^Sheet \d+:\s*', '', raw_title)[:28].strip()
        if not tab_name:
            tab_name = f"Sheet {snum}"

        ws = wb.create_sheet(tab_name)
        ws.sheet_properties.tabColor = tab_colors.get(snum, '2E75B6')
        ws.sheet_view.showGridLines = False

        if sd['tables']:
            write_sheet(ws, sd)
        else:
            ws.cell(row=1, column=1, value=f"[Sheet {snum}: tidak ada tabel]")

        total_rows = sum(len(t['rows']) for t in sd['tables'])
        print(f"  -> Tab '{tab_name}': {total_rows} baris")

    try:
        wb.save(DEST_XLSX)
        print(f"\n[SUKSES] Excel berhasil digenerate:\n  -> {DEST_XLSX}")
        return True
    except PermissionError:
        print(f"\n[PERINGATAN] File Excel '{os.path.basename(DEST_XLSX)}' sedang dibuka di aplikasi lain.")
        return False
    except Exception as e:
        print(f"\n[ERROR] Gagal menyimpan file: {e}")
        return False

if __name__ == '__main__':
    convert()
