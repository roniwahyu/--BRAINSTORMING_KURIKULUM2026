"""
=============================================================================
COMPREHENSIVE EXCEL EXPORTER — KURIKULUM OBE SISTEKIN 2026
=============================================================================
Script ini mengonversi dokumen-dokumen Markdown kurikulum ke dalam format Excel (.xlsx):
1. 001_ANALISIS_VMTS_DAN_POSITIONING_STRATEGIS_SISTEKIN.xlsx
2. 011_IMPLEMENTASI_OBE_SISTEKIN2026_TABLES.xlsx
3. MASTER_KURIKULUM_OBE_SISTEKIN_2026.xlsx (Workbook Lengkap Multi-Dokumen)
=============================================================================
"""

import os
import sys
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if not isinstance(sys.stdout, io.TextIOWrapper) or sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKDIR = os.path.dirname(SCRIPT_DIR)

COLOR = {
    'sheet_bg': '1F3864',
    'sheet_fg': 'FFFFFF',
    'col_header': '2E75B6',
    'col_fg': 'FFFFFF',
    'section': 'BDD7EE',
    'section_fg': '1F3864',
    'alt_row': 'EEF4FB',
    'border': 'B8CCE4',
    'note': 'FFF2CC',
}

thin = Side(style='thin', color=COLOR['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style_header(cell, fg='FFFFFF', bg='2E75B6', bold=True, sz=10, wrap=True):
    cell.font = Font(bold=bold, color=fg, size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='center', horizontal='center')
    cell.border = BORDER

def style_data(cell, bg='FFFFFF', bold=False, sz=10, halign='left', wrap=True):
    cell.font = Font(bold=bold, color='000000', size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='top', horizontal=halign)
    cell.border = BORDER

def style_section(cell):
    cell.font = Font(bold=True, color=COLOR['section_fg'], size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['section'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_note(cell):
    cell.font = Font(bold=True, color='7F6000', size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['note'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_title(ws, title, subtitle="", row_offset=1):
    ws.row_dimensions[row_offset].height = 24
    c1 = ws.cell(row=row_offset, column=1, value=title)
    c1.font = Font(bold=True, color=COLOR['sheet_fg'], size=13, name='Calibri')
    c1.fill = cell_fill(COLOR['sheet_bg'])
    c1.alignment = Alignment(vertical='center', horizontal='left')

    if subtitle:
        ws.row_dimensions[row_offset + 1].height = 16
        c2 = ws.cell(row=row_offset + 1, column=1, value=subtitle)
        c2.font = Font(bold=False, color=COLOR['sheet_fg'], size=10, name='Calibri')
        c2.fill = cell_fill(COLOR['sheet_bg'])
        c2.alignment = Alignment(vertical='center', horizontal='left')

def extract_tables_from_md(md_text):
    """Mengekstrak seluruh tabel Markdown dari teks."""
    tables = []
    # Cari section header terdekat sebelum tabel
    lines = md_text.split('\n')
    current_section = "Tabel Data"
    
    table_lines = []
    in_table = False
    
    for line in lines:
        if line.startswith('#'):
            current_section = re.sub(r'^#+\s*', '', line).strip()
        
        if re.match(r'^\s*\|.+\|\s*$', line):
            in_table = True
            table_lines.append(line.strip())
        else:
            if in_table:
                # Selesai 1 blok tabel
                if len(table_lines) >= 2:
                    rows = []
                    for t_line in table_lines:
                        if re.match(r'^\s*\|[-| :]+\|\s*$', t_line):
                            continue
                        cells = [c.strip() for c in t_line.strip('|').split('|')]
                        rows.append(cells)
                    if rows:
                        tables.append({'section': current_section, 'rows': rows})
                table_lines = []
                in_table = False
                
    if in_table and len(table_lines) >= 2:
        rows = []
        for t_line in table_lines:
            if re.match(r'^\s*\|[-| :]+\|\s*$', t_line):
                continue
            cells = [c.strip() for c in t_line.strip('|').split('|')]
            rows.append(cells)
        if rows:
            tables.append({'section': current_section, 'rows': rows})
            
    return tables

def write_md_tables_to_sheet(ws, title, subtitle, tables):
    style_title(ws, title, subtitle, row_offset=1)
    ws.freeze_panes = 'A4'
    current_row = 3
    
    for tbl_idx, tbl in enumerate(tables):
        rows = tbl['rows']
        if not rows:
            continue
            
        header = rows[0]
        n_cols = len(header)
        
        if tbl_idx == 0:
            try:
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(n_cols, 6))
                if subtitle:
                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(n_cols, 6))
            except Exception:
                pass
                
        # Tulis label section jika ada lebih dari 1 tabel
        if len(tables) > 1 and tbl.get('section'):
            ws.row_dimensions[current_row].height = 20
            c_sec = ws.cell(row=current_row, column=1, value=tbl['section'])
            style_section(c_sec)
            try:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max(n_cols, 6))
            except Exception:
                pass
            current_row += 1
            
        # Header tabel
        ws.row_dimensions[current_row].height = 26
        for col_i, val in enumerate(header, start=1):
            c = ws.cell(row=current_row, column=col_i, value=val)
            style_header(c, sz=10)
        current_row += 1
        
        # Data rows
        for row_i, row in enumerate(rows[1:], start=1):
            row = (row + [''] * n_cols)[:n_cols]
            bg = COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF'
            
            # Cek jika baris note / total
            is_note = bool(re.match(r'^(Total|Jumlah|Validasi|Ringkasan|Catatan|Strategi)', row[0], re.IGNORECASE))
            styler = style_note if is_note else None
            if is_note:
                bg = COLOR['note']
                
            ws.row_dimensions[current_row].height = 16
            for col_i, val in enumerate(row, start=1):
                c = ws.cell(row=current_row, column=col_i, value=val)
                if styler:
                    styler(c)
                else:
                    halign = 'center' if (col_i <= 4 and len(str(val)) <= 8) else 'left'
                    style_data(c, bg=bg, bold=False, halign=halign)
            current_row += 1
            
        current_row += 1
        
    # Auto column width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, max_len * 1.15)
        
    ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width or 10, 8)

def convert_single_file(md_filename):
    """Mengonversi satu file .md menjadi file .xlsx tersendiri."""
    src_path = os.path.join(WORKDIR, md_filename)
    if not os.path.exists(src_path):
        print(f"[SKIP] File tidak ditemukan: {md_filename}")
        return None
        
    out_xlsx = os.path.join(WORKDIR, md_filename.replace('.md', '.xlsx'))
    with open(src_path, 'r', encoding='utf-8') as f:
        md_text = f.read()
        
    tables = extract_tables_from_md(md_text)
    if not tables:
        print(f"[INFO] Tidak ada tabel pada: {md_filename}")
        return None
        
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    ws = wb.create_sheet("Data")
    ws.sheet_properties.tabColor = "1F3864"
    ws.sheet_view.showGridLines = False
    
    title = f"SISTEKIN 2026 — {md_filename.split('.')[0]}"
    write_md_tables_to_sheet(ws, title, "Dokumen Kurikulum KPT-OBE SISTEKIN 2026", tables)
    
    try:
        wb.save(out_xlsx)
        print(f"  -> Generated: {os.path.basename(out_xlsx)} ({os.path.getsize(out_xlsx)} bytes)")
        return out_xlsx
    except PermissionError:
        print(f"[PERINGATAN] File {out_xlsx} sedang dibuka di aplikasi lain.")
        return None

def convert_master_workbook():
    """Mengonversi seluruh dokumen kurikulum ke dalam satu Master Workbook Excel."""
    master_path = os.path.join(WORKDIR, "MASTER_KURIKULUM_OBE_SISTEKIN_2026.xlsx")
    print(f"\n[INFO] Menyusun Master Workbook Multi-Sheet:")
    print(f"  -> {master_path}")
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Cover Sheet
    cover = wb.create_sheet("Cover & Summary")
    cover.sheet_properties.tabColor = "1F3864"
    cover.column_dimensions['A'].width = 5
    cover.column_dimensions['B'].width = 85
    
    cover_rows = [
        (1, "MASTER WORKBOOK KURIKULUM OBE", True, 20, '1F3864', 'FFFFFF'),
        (2, "S1 SISTEM DAN TEKNOLOGI INFORMASI 2026", True, 14, '1F3864', 'FFFFFF'),
        (3, "Fakultas Sains dan Teknologi Informasi (FSTI) — Universitas Widyagama Malang", False, 12, '2E75B6', 'FFFFFF'),
        (4, "", False, 10, 'FFFFFF', '000000'),
        (5, "Versi: Revisi Final Definitif 2026 (Kurikulum KPT-OBE)", False, 11, 'EEF4FB', '1F3864'),
        (6, "Standar: Permendikbudristek No. 53/2023 | APTIKOM IS2020 & IT2017 | LAM INFOKOM", False, 11, 'EEF4FB', '1F3864'),
        (7, "Beban Paket Ditempuh: 55 MK / 146 SKS (Portofolio Ditawarkan: 67 MK / 182 SKS)", False, 11, 'EEF4FB', '1F3864'),
        (8, "", False, 10, 'FFFFFF', '000000'),
        (9, "Konsensus Ground Truth Final Terverifikasi:", True, 11, 'FFF2CC', '7F6000'),
        (10, "  ✓  VMTS 2045: Smart Systems, Applied AI Integration, dan Technopreneurship", False, 11, 'FFF2CC', '7F6000'),
        (11, "  ✓  4 Profil Lulusan (PL-1 s.d. PL-4) & 3 PEO Terukur 3-5 Tahun", False, 11, 'FFF2CC', '7F6000'),
        (12, "  ✓  14 CPL Terstandar (S1, KU1-3, P1-4, KK1-6 @2 per peminatan)", False, 11, 'FFF2CC', '7F6000'),
        (13, "  ✓  3 Peminatan Seimbang @ 18 SKS / 6 MK (P1 Smart Sys, P2 Cloud/Cyber, P3 Platform)", False, 11, 'FFF2CC', '7F6000'),
        (14, "  ✓  STI-405 Dasar Keamanan Informasi = 2 SKS & STI-602 Smart City = 2 SKS", False, 11, 'FFF2CC', '7F6000'),
        (15, "  ✓  MKU-406 Agama II & MKU-508 Kewirausahaan II = 0 SKS (Kebijakan Khusus UWG)", False, 11, 'FFF2CC', '7F6000'),
        (16, "  ✓  Skema 4x Asesmen Baku per MK (Tugas 1 [20%], UTS [25-30%], Tugas 2 [20-25%], UAS [30%])", False, 11, 'FFF2CC', '7F6000'),
    ]
    
    for r, val, bold, sz, bg, fg in cover_rows:
        cover.row_dimensions[r].height = sz * 1.8
        c = cover.cell(row=r, column=2, value=val)
        c.font = Font(bold=bold, size=sz, color=fg, name='Calibri')
        c.fill = cell_fill(bg)
        c.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        ca = cover.cell(row=r, column=1, value='')
        ca.fill = cell_fill(bg)

    # Tambahkan Sheet dari masing-masing Dokumen Utama
    doc_sources = [
        ("001_ANALISIS_VMTS_DAN_POSITIONING_STRATEGIS_SISTEKIN.md", "001 VMTS & SWOT", "2E75B6"),
        ("002_FORMULASI_3_PEO_DAN_4_PROFIL_LULUSAN_SISTEKIN.md", "002 PEO & 4 PL", "ED7D31"),
        ("003_STANDAR_14_CPL_DAN_PEMETAAN_BoK_APTIKOM.md", "003 14 CPL & BoK", "4472C4"),
        ("004_MATRIKS_KETERLACAKAN_OBE_VMTS_PEO_PL_CPL_MK.md", "004 Matriks Makro OBE", "5B9BD5"),
        ("005_STRUKTUR_KURIKULUM_8_SEMESTER_DAN_PEMINATAN.md", "005 Struktur 8 Semester", "70AD47"),
        ("006_DISTRIBUSI_DAN_PANDUAN_MK_PEMINATAN_MBKM.md", "006 3 Peminatan & MBKM", "10B981"),
        ("008_SISTEM_ASESMEN_OBE_FORMULA_CPL_DAN_RUBRIK_MASTER.md", "008 Asesmen & Rubrik", "FFC000"),
        ("009E_RINGKASAN_CPL_LENGKAP_PEMETAAN_BoK.md", "009E Matriks Kompilasi CPL", "7030A0"),
        ("010_INSTRUMEN_TRACER_STUDY_DAN_EVALUASI_PEO_PPEPP.md", "010 Tracer Study & PPEPP", "EC4899"),
    ]
    
    for filename, tab_label, tab_col in doc_sources:
        src_path = os.path.join(WORKDIR, filename)
        if not os.path.exists(src_path):
            continue
            
        with open(src_path, 'r', encoding='utf-8') as f:
            md_text = f.read()
            
        tables = extract_tables_from_md(md_text)
        if not tables:
            continue
            
        ws = wb.create_sheet(tab_label)
        ws.sheet_properties.tabColor = tab_col
        ws.sheet_view.showGridLines = False
        
        doc_title = f"SISTEKIN 2026 — {tab_label}"
        write_md_tables_to_sheet(ws, doc_title, "Dokumen Kurikulum KPT-OBE SISTEKIN 2026", tables)
        total_rows = sum(len(t['rows']) for t in tables)
        print(f"  -> Tab '{tab_label}': {len(tables)} tabel, {total_rows} baris")
        
    try:
        wb.save(master_path)
        print(f"\n[SUKSES] Master Workbook berhasil digenerate:\n  -> {master_path}")
    except PermissionError:
        print(f"[PERINGATAN] File {master_path} sedang dibuka di aplikasi lain.")

def main():
    print("=====================================================================")
    print("  EXCEL EXPORTER: KURIKULUM OBE SISTEKIN 2026")
    print("=====================================================================")
    
    docs_to_convert = [
        "001_ANALISIS_VMTS_DAN_POSITIONING_STRATEGIS_SISTEKIN.md",
        "002_FORMULASI_3_PEO_DAN_4_PROFIL_LULUSAN_SISTEKIN.md",
        "003_STANDAR_14_CPL_DAN_PEMETAAN_BoK_APTIKOM.md",
        "004_MATRIKS_KETERLACAKAN_OBE_VMTS_PEO_PL_CPL_MK.md",
        "005_STRUKTUR_KURIKULUM_8_SEMESTER_DAN_PEMINATAN.md",
        "006_DISTRIBUSI_DAN_PANDUAN_MK_PEMINATAN_MBKM.md",
        "007_FORMULASI_CPMK_DAN_SUB_CPMK_PORTFOLIO_LENGKAP.md",
        "008_SISTEM_ASESMEN_OBE_FORMULA_CPL_DAN_RUBRIK_MASTER.md",
        "009_LANGKAH2_CPL_FORMAL.md",
        "009A_CPL_SIKAP_SISTEKIN.md",
        "009B_CPL_KETERAMPILAN_UMUM_SISTEKIN.md",
        "009C_CPL_PENGETAHUAN_SISTEKIN.md",
        "009D_CPL_KETERAMPILAN_KHUSUS_SISTEKIN.md",
        "009E_RINGKASAN_CPL_LENGKAP_PEMETAAN_BoK.md",
        "009_PEDOMAN_CAPSTONE_PROJECT_DAN_TUGAS_AKHIR_NON_SKRIPSI.md",
        "010_INSTRUMEN_TRACER_STUDY_DAN_EVALUASI_PEO_PPEPP.md",
    ]
    
    print("\n1. Mengonversi Masing-Masing Dokumen Kurikulum ke Excel Individu:")
    for md_doc in docs_to_convert:
        convert_single_file(md_doc)
        
    print("\n2. Mengonversi Dokumen 011 (14 Sheet) ke Excel:")
    try:
        import export_011_tables_to_excel
        export_011_tables_to_excel.convert()
    except Exception as e:
        print(f"  [Catatan 011]: {e}")
        
    print("\n3. Menyusun Master Workbook Lengkap (MASTER_KURIKULUM_OBE_SISTEKIN_2026.xlsx):")
    try:
        convert_master_workbook()
    except Exception as e:
        print(f"  [Catatan Master]: {e}")
        
    print("\n[SELESAI] Proses ekspor Excel selesai!")

if __name__ == '__main__':
    main()
