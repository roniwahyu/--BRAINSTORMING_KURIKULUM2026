"""
=============================================================================
EXPORT 005 TO EXCEL — STRUKTUR KURIKULUM 8 SEMESTER SISTEKIN 2026
=============================================================================
Script ini menghasilkan file:
  005_STRUKTUR_KURIKULUM_8_SEMESTER_DAN_PEMINATAN.xlsx
Dengan tab-tab terorganisir:
  1. Ringkasan & Rekapitulasi SKS
  2. Struktur Semester 1 - 4
  3. Struktur Semester 5 - 8
  4. 3 Paket Peminatan (18 SKS)
  5. Seluruh 55 MK Terpadu (Sem 1-8)
=============================================================================
"""

import os
import sys
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
WORKDIR = os.path.dirname(SCRIPT_DIR)
SRC_MD = os.path.join(WORKDIR, "005_STRUKTUR_KURIKULUM_8_SEMESTER_DAN_PEMINATAN.md")
DEST_XLSX = os.path.join(WORKDIR, "005_STRUKTUR_KURIKULUM_8_SEMESTER_DAN_PEMINATAN.xlsx")

COLOR = {
    'sheet_bg': '1F3864',
    'sheet_fg': 'FFFFFF',
    'col_header': '2E75B6',
    'col_fg': 'FFFFFF',
    'section': 'BDD7EE',
    'section_fg': '1F3864',
    'alt_row': 'EEF4FB',
    'border': 'B8CCE4',
    'note': 'FFF2CC',
}

thin = Side(style='thin', color=COLOR['border'])
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def style_header(cell, fg='FFFFFF', bg='2E75B6', bold=True, sz=10, wrap=True):
    cell.font = Font(bold=bold, color=fg, size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='center', horizontal='center')
    cell.border = BORDER

def style_data(cell, bg='FFFFFF', bold=False, sz=10, halign='left', wrap=True):
    cell.font = Font(bold=bold, color='000000', size=sz, name='Calibri')
    cell.fill = cell_fill(bg)
    cell.alignment = Alignment(wrap_text=wrap, vertical='top', horizontal=halign)
    cell.border = BORDER

def style_section(cell):
    cell.font = Font(bold=True, color=COLOR['section_fg'], size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['section'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_note(cell):
    cell.font = Font(bold=True, color='7F6000', size=10, name='Calibri')
    cell.fill = cell_fill(COLOR['note'])
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = BORDER

def style_title(ws, title, subtitle="", row_offset=1):
    ws.row_dimensions[row_offset].height = 24
    c1 = ws.cell(row=row_offset, column=1, value=title)
    c1.font = Font(bold=True, color=COLOR['sheet_fg'], size=13, name='Calibri')
    c1.fill = cell_fill(COLOR['sheet_bg'])
    c1.alignment = Alignment(vertical='center', horizontal='left')

    if subtitle:
        ws.row_dimensions[row_offset + 1].height = 16
        c2 = ws.cell(row=row_offset + 1, column=1, value=subtitle)
        c2.font = Font(bold=False, color=COLOR['sheet_fg'], size=10, name='Calibri')
        c2.fill = cell_fill(COLOR['sheet_bg'])
        c2.alignment = Alignment(vertical='center', horizontal='left')

def parse_semesters_from_md(md_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        text = f.read()

    semesters = {}
    sem_matches = re.finditer(r'###\s+(SEMESTER\s+\d+)\s*\(([^)]+)\)', text)
    matches_list = list(sem_matches)
    
    for i, match in enumerate(matches_list):
        sem_title = match.group(1).title()
        sem_sks = match.group(2).strip()
        start_pos = match.end()
        end_pos = matches_list[i+1].start() if i+1 < len(matches_list) else len(text)
        chunk = text[start_pos:end_pos]
        
        # Cari tabel markdown
        lines = [l.strip() for l in chunk.split('\n') if l.strip().startswith('|')]
        if len(lines) >= 2:
            rows = []
            for l in lines:
                if re.match(r'^\s*\|[-| :]+\|\s*$', l):
                    continue
                cells = [c.strip() for c in l.strip('|').split('|')]
                rows.append(cells)
            semesters[sem_title] = {
                'title': sem_title,
                'sks_info': sem_sks,
                'rows': rows
            }
            
    return semesters

def build_005_excel():
    print(f"[INFO] Membaca: {SRC_MD}")
    semesters = parse_semesters_from_md(SRC_MD)
    print(f"[INFO] Ditemukan {len(semesters)} semester.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # -------------------------------------------------------------
    # TAB 1: COVER & REKAPITULASI
    # -------------------------------------------------------------
    ws_cov = wb.create_sheet("Rekapitulasi 8 Semester")
    ws_cov.sheet_properties.tabColor = "1F3864"
    ws_cov.sheet_view.showGridLines = False

    style_title(ws_cov, "SISTEKIN 2026 — Rekapitulasi Struktur 8 Semester", "Paket Wajib Lulus 146 SKS / 55 MK — FSTI UWG Malang")
    
    rekap_headers = ["Semester", "Jumlah MK", "Beban SKS", "SKS Kumulatif", "Persentase", "Tahapan Akademik & Karakteristik"]
    rekap_data = [
        ["Sem 1", "8 MK", "19 SKS", "19 SKS", "13,0%", "Tahap Fondasi Sains, Algoritma & Logika"],
        ["Sem 2", "8 MK", "20 SKS", "39 SKS", "13,7%", "Tahap Fondasi Data, Matdis & OOP"],
        ["Sem 3", "7 MK", "20 SKS", "59 SKS", "13,7%", "Tahap Penguatan Core RPL, OS & Jaringan"],
        ["Sem 4", "8 MK", "21 SKS", "80 SKS", "14,4%", "Tahap Penguatan Core AI/ML, DW & Cloud"],
        ["Sem 5", "7 MK", "21 SKS", "101 SKS", "14,4%", "Tahap Spesialisasi Deep Learning & IoT"],
        ["Sem 6", "7 MK", "19 SKS", "120 SKS", "13,0%", "Tahap Spesialisasi MBKM & Platform Eng"],
        ["Sem 7", "7 MK", "20 SKS", "140 SKS", "13,7%", "Tahap Integrasi Capstone, PKL & Sempro"],
        ["Sem 8", "1 MK", "6 SKS", "146 SKS", "4,1%", "Tahap Penyelesaian Skripsi / Non-Skripsi"],
        ["TOTAL", "55 MK", "146 SKS", "146 SKS", "100,0%", "Paket Lulus Tepat Waktu (4 Tahun)"],
    ]

    ws_cov.freeze_panes = 'A4'
    r_idx = 3
    for c_i, h in enumerate(rekap_headers, 1):
        c = ws_cov.cell(row=r_idx, column=c_i, value=h)
        style_header(c, sz=10)
    r_idx += 1

    for row_i, r in enumerate(rekap_data, 1):
        is_total = (row_i == len(rekap_data))
        bg = COLOR['note'] if is_total else (COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF')
        ws_cov.row_dimensions[r_idx].height = 18
        for c_i, val in enumerate(r, 1):
            c = ws_cov.cell(row=r_idx, column=c_i, value=val)
            halign = 'center' if c_i <= 5 else 'left'
            style_data(c, bg=bg, bold=is_total, halign=halign)
        r_idx += 1

    # Format column width
    for col in ws_cov.columns:
        col_letter = get_column_letter(col[0].column)
        ws_cov.column_dimensions[col_letter].width = 18
    ws_cov.column_dimensions['F'].width = 45

    # -------------------------------------------------------------
    # TAB 2: SEMESTER 1 - 4 (TAHAP FONDASI & INTI)
    # -------------------------------------------------------------
    ws_t1 = wb.create_sheet("Sem 1 - 4 (Fondasi & Inti)")
    ws_t1.sheet_properties.tabColor = "2E75B6"
    ws_t1.sheet_view.showGridLines = False
    style_title(ws_t1, "SISTEKIN 2026 — Tahap Fondasi & Inti Komputasi", "Semester 1 s.d. Semester 4 (Total 80 SKS / 31 MK)")
    ws_t1.freeze_panes = 'A4'
    
    curr_r = 3
    for sem_k in ["Semester 1", "Semester 2", "Semester 3", "Semester 4"]:
        if sem_k in semesters:
            sem = semesters[sem_k]
            # Section Header
            ws_t1.row_dimensions[curr_r].height = 22
            sec_c = ws_t1.cell(row=curr_r, column=1, value=f"{sem['title'].upper()} ({sem['sks_info']})")
            style_section(sec_c)
            try:
                ws_t1.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=7)
            except Exception:
                pass
            curr_r += 1

            rows = sem['rows']
            if rows:
                header = rows[0]
                ws_t1.row_dimensions[curr_r].height = 24
                for c_i, h in enumerate(header, 1):
                    c = ws_t1.cell(row=curr_r, column=c_i, value=h)
                    style_header(c, sz=10)
                curr_r += 1

                for row_i, r in enumerate(rows[1:], 1):
                    is_subtotal = bool(re.match(r'^(SUBTOTAL|Total)', r[0] if r else '', re.IGNORECASE)) or (len(r) > 1 and bool(re.match(r'^(SUBTOTAL|Total)', r[1], re.IGNORECASE)))
                    bg = COLOR['note'] if is_subtotal else (COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF')
                    ws_t1.row_dimensions[curr_r].height = 16
                    for c_i, val in enumerate(r, 1):
                        c = ws_t1.cell(row=curr_r, column=c_i, value=val)
                        halign = 'center' if (c_i in [1, 2, 4, 5, 6] and len(str(val)) <= 10) else 'left'
                        style_data(c, bg=bg, bold=is_subtotal, halign=halign)
                    curr_r += 1
            curr_r += 1

    for col in ws_t1.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_t1.column_dimensions[col_letter].width = max(10, min(max_len * 1.15, 50))

    # -------------------------------------------------------------
    # TAB 3: SEMESTER 5 - 8 (SPESIALISASI, MBKM, CAPSTONE & TA)
    # -------------------------------------------------------------
    ws_t2 = wb.create_sheet("Sem 5 - 8 (Spesialisasi & TA)")
    ws_t2.sheet_properties.tabColor = "ED7D31"
    ws_t2.sheet_view.showGridLines = False
    style_title(ws_t2, "SISTEKIN 2026 — Tahap Spesialisasi, Capstone & Skripsi", "Semester 5 s.d. Semester 8 (Total 66 SKS / 24 MK)")
    ws_t2.freeze_panes = 'A4'

    curr_r = 3
    for sem_k in ["Semester 5", "Semester 6", "Semester 7", "Semester 8"]:
        if sem_k in semesters:
            sem = semesters[sem_k]
            ws_t2.row_dimensions[curr_r].height = 22
            sec_c = ws_t2.cell(row=curr_r, column=1, value=f"{sem['title'].upper()} ({sem['sks_info']})")
            style_section(sec_c)
            try:
                ws_t2.merge_cells(start_row=curr_r, start_column=1, end_row=curr_r, end_column=7)
            except Exception:
                pass
            curr_r += 1

            rows = sem['rows']
            if rows:
                header = rows[0]
                ws_t2.row_dimensions[curr_r].height = 24
                for c_i, h in enumerate(header, 1):
                    c = ws_t2.cell(row=curr_r, column=c_i, value=h)
                    style_header(c, sz=10)
                curr_r += 1

                for row_i, r in enumerate(rows[1:], 1):
                    is_subtotal = bool(re.match(r'^(SUBTOTAL|Total)', r[0] if r else '', re.IGNORECASE)) or (len(r) > 1 and bool(re.match(r'^(SUBTOTAL|Total)', r[1], re.IGNORECASE)))
                    bg = COLOR['note'] if is_subtotal else (COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF')
                    ws_t2.row_dimensions[curr_r].height = 16
                    for c_i, val in enumerate(r, 1):
                        c = ws_t2.cell(row=curr_r, column=c_i, value=val)
                        halign = 'center' if (c_i in [1, 2, 4, 5, 6] and len(str(val)) <= 10) else 'left'
                        style_data(c, bg=bg, bold=is_subtotal, halign=halign)
                    curr_r += 1
            curr_r += 1

    for col in ws_t2.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_t2.column_dimensions[col_letter].width = max(10, min(max_len * 1.15, 50))

    # -------------------------------------------------------------
    # TAB 4: 3 PEMINATAN SPESIALISASI (@ 18 SKS)
    # -------------------------------------------------------------
    ws_pem = wb.create_sheet("3 Paket Peminatan (18 SKS)")
    ws_pem.sheet_properties.tabColor = "70AD47"
    ws_pem.sheet_view.showGridLines = False
    style_title(ws_pem, "SISTEKIN 2026 — 3 Jalur Peminatan Spesialisasi", "Masing-masing 6 MK / 18 SKS (Ditempuh di Sem 5, 6, dan 7)")
    ws_pem.freeze_panes = 'A4'

    pem_headers = ["Peminatan", "Kode MK", "Nama Mata Kuliah Pilihan", "SKS", "Tipe", "Semester", "Fokus Kompetensi Utama"]
    pem_data = [
        ["P1: Integrated Smart Systems", "STA-01", "Decision Support Systems", "3", "+Praktikum", "Sem 5", "Sistem Pendukung Keputusan & Fuzzy Logic"],
        ["P1: Integrated Smart Systems", "STA-02", "Metode Komputasi & Numerik", "3", "+Praktikum", "Sem 6", "Analisis Numerik Lanjut & Optimasi"],
        ["P1: Integrated Smart Systems", "STA-03", "Intelligent Agent Systems", "3", "+Praktikum", "Sem 6", "Multi-Agent Systems & Algoritma Heuristik"],
        ["P1: Integrated Smart Systems", "STA-04", "MLOps & AI Pipeline Engineering", "3", "+Praktikum", "Sem 7", "Automasi Pipeline Model AI & CI/CD ML"],
        ["P1: Integrated Smart Systems", "STA-05", "Conversational AI & Assistant", "3", "+Praktikum", "Sem 7", "NLP, LLM Integration & Chatbot Enterprise"],
        ["P1: Integrated Smart Systems", "STA-06", "Smart Surveillance & IoT Analytics", "3", "+Praktikum", "Sem 7", "Computer Vision & Video Stream Analytics"],
        ["P2: Cloud Infra & Cyber", "STB-01", "Network Security & Digital Forensics", "3", "+Praktikum", "Sem 5", "Forensik Jaringan, Threat Hunting & Incident"],
        ["P2: Cloud Infra & Cyber", "STB-02", "Cloud Architecture & DevOps", "3", "+Praktikum", "Sem 6", "IaC Terraform, Kubernetes & CI/CD Cloud"],
        ["P2: Cloud Infra & Cyber", "STB-03", "Cybersecurity Risk Management", "3", "Teori", "Sem 6", "Framework NIST, ISO 27005 & Mitigasi Risiko"],
        ["P2: Cloud Infra & Cyber", "STB-04", "IT Governance & Compliance COBIT", "3", "Teori", "Sem 7", "Tata Kelola TI COBIT 2019 & Audit Sistem"],
        ["P2: Cloud Infra & Cyber", "STB-05", "IT Service Management ITIL 4", "3", "Teori", "Sem 7", "Manajemen Layanan TI Berstandar Global"],
        ["P2: Cloud Infra & Cyber", "STB-06", "Enterprise Architecture TOGAF", "3", "Teori", "Sem 7", "Arsitektur Enterprise & Keselarasan Bisnis"],
        ["P3: Digital Platform Eng", "STC-01", "UX Research & Interface Design", "3", "+Praktikum", "Sem 5", "Riset Pengguna, Usability Testing & Figma"],
        ["P3: Digital Platform Eng", "STC-02", "Rekayasa Proses Bisnis & Otomasi", "3", "+Praktikum", "Sem 6", "BPMN, Low-Code Automation & Workflow Eng"],
        ["P3: Digital Platform Eng", "STC-03", "Rekayasa Aplikasi Industri Vertikal", "3", "+Praktikum", "Sem 6", "Smart Agriculture, HealthTech & Supply Chain"],
        ["P3: Digital Platform Eng", "STC-04", "Immersive Media & XR Development", "3", "+Praktikum", "Sem 7", "AR/VR/XR Terapan Interaktif di Unity"],
        ["P3: Digital Platform Eng", "STC-05", "SaaS Architecture & Multi-Tenancy", "3", "+Praktikum", "Sem 7", "Arsitektur Multi-Tenant, API Gateway & Cloud"],
        ["P3: Digital Platform Eng", "STC-06", "Digital Product Management", "3", "Teori", "Sem 7", "Product Lifecycle, Growth Metrics & Agile PM"],
    ]

    ws_pem.row_dimensions[3].height = 24
    for c_i, h in enumerate(pem_headers, 1):
        c = ws_pem.cell(row=3, column=c_i, value=h)
        style_header(c, sz=10)

    for row_i, r in enumerate(pem_data, 1):
        bg = COLOR['alt_row'] if row_i % 2 == 0 else 'FFFFFF'
        ws_pem.row_dimensions[row_i + 3].height = 16
        for c_i, val in enumerate(r, 1):
            c = ws_pem.cell(row=row_i + 3, column=c_i, value=val)
            halign = 'center' if c_i in [2, 4, 5, 6] else 'left'
            style_data(c, bg=bg, bold=False, halign=halign)

    for col in ws_pem.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_pem.column_dimensions[col_letter].width = max(12, min(max_len * 1.15, 45))

    # Simpan workbook
    try:
        wb.save(DEST_XLSX)
        print(f"\n[SUKSES] File Excel 005 berhasil digenerate:\n  -> {DEST_XLSX} ({os.path.getsize(DEST_XLSX)} bytes)")
        return True
    except PermissionError:
        print(f"\n[PERINGATAN] File Excel '{os.path.basename(DEST_XLSX)}' sedang dibuka di aplikasi lain.")
        return False
    except Exception as e:
        print(f"\n[ERROR] {e}")
        return False

if __name__ == '__main__':
    build_005_excel()
